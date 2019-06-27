#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2019/6/17 12:24
# @Author  : Jandy
# @Email   : jandyqin@gmail.com
# @File    ：do_excel.py
# @Software: PyCharm

from openpyxl import load_workbook

from common.test_base import TestBase
from common.read_config import config
import openpyxl
import os
from common.constant import data_path


class DoExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.wb = load_workbook(file_path)
        self.sheet = self.wb[sheet_name]

    def __del__(self):
        self.wb.close

    def read_excel(self):
        test_data = []
        for i in range(2, self.sheet.max_row + 1):
            sub_data = {}
            sub_data['run'] = self.sheet.cell(row=i, column=1).value
            sub_data['desc'] = self.sheet.cell(row=i, column=2).value
            sub_data['method'] = self.sheet.cell(row=i, column=3).value
            sub_data['url'] = self.sheet.cell(row=i, column=4).value
            sub_data['param'] = self.sheet.cell(row=i, column=5).value
            sub_data['verify'] = self.sheet.cell(row=i, column=6).value
            sub_data['save'] = self.sheet.cell(row=i, column=6).value
            test_data.append(sub_data)
        return test_data

    def write_back(self, row, column, value):
        '''
        写入数据
        :param row: 写入的行数
        :param column: 写入的列
        :param value: 写入的值
        :return:
        '''

        # 为指定的行列写入指定的值
        self.sheet.cell(row=row, column=column).value = value
        # 保存
        self.wb.save(self.file_path)


class Case:
    '''
    用来存放用例数据
    '''

    def __init__(self, attr):
        '''

        :param attr:每行的用例数据，如：[('data', "('xyc1234','1234567','1234567')"), ('expected', '{"code": 1, "msg": "注册成功"}')]
        '''
        for case in attr:
            # if case[0] != None:
            setattr(self, case[0], case[1])


class ReadExcel:
    """
    用来读取excel数据
    """

    def __init__(self, filename, sheetname):
        """
        接收文件名，表单名
        初始化加载工作簿，表单
        :param filename:  --文件名，str
        :param sheetname: --表单名，str
        """
        self.filename = filename
        # 打开工作簿1
        self.wb = openpyxl.load_workbook(filename)
        # 选定工作表
        self.sh = self.wb[sheetname]

    def __del__(self):
        '''
        对象销毁后，关闭工作簿
        :return:
        '''
        self.wb.close()

    def read_data_obj(self):
        '''
        按行读取数据，表单所有数据
        每个用例存储在一个对象中
        :return: 返回一个列表，列表中每个元素为一个用例对象
        '''
        # 将每行的数据转成列表
        row_data = list(self.sh.rows)
        # 获取表单的表头信息
        titles = []
        titles.append('case_row_id')
        for title in row_data[0]:
            if title.value != None:
                titles.append(title.value)

        # 定义一个空列表用来存储所有的用例
        cases = []
        # 遍历出所有测试用例
        i = 2
        for case in row_data[1:]:
            # data用来临时存放用例数据
            data = []

            data.append(i)
            i+=1
            for cell in case:
                data.append(cell.value)
            # 将该条数据放入cases中
            data = list(zip(titles, data))
            case_obj = Case(data)
            cases.append(case_obj)
        return cases

    def r_data_obj(self, list1):
        '''
        # list1 参数为一个列表，传入的是指定读取数据的列,比如[1,2,3]、
        # 每一行[1,3,5]列的数据，读取出来就作为一条测试用例，放在对象中属性中
        # 所有的用例对象放在列表中并且进行返回
        :param list1: 必须是列表，列表中只能是正整数，如：[1,3,5]
        :return: 返回一个列表，列表里的每个对象代表一条用例
        '''

        list1 = eval(list1)
        # 列表为空，读取所有数据
        if list1 == []:
            return self.read_data_obj()
        # 读取指定列的标题
        titles = []
        for x in list1:
            title = self.sh.cell(row=1, column=x).value
            titles.append(title)
        # 用来存储所有用例
        cases = []
        # 获取最大行数，然后从第二行遍历
        for x in range(2, self.sh.max_row + 1):
            # case 临时存储用例数据
            case = []
            # 遍历list1的所有元素
            for y in list1:
                cell_value = self.sh.cell(row=x, column=y).value
                case.append(cell_value)
            # 将每条用例和对应的列名打包成列表
            data = list(zip(titles, case))
            # 将每条用例数据传到Case类里设置属性
            case_obj = Case(data)
            # 将每条用例的对象追加到cases列表里
            cases.append(case_obj)
        return cases

    def write(self, row, column, msg):
        '''

        :param row: 行 -->int
        :param column: 列-->int
        :param msg:   写入数据
        :return:
        '''
        self.sh.cell(row=row, column=column, value=msg)
        self.wb.save(self.filename)


# 文件名 ，表单名
file_name = config.get('excel', 'excel_name')
file_path = os.path.join(data_path, file_name)
sheet_name = config.get('excel', 'sheet_name')
# file_path = 'E:/PycharmProjects/auto_test/test_data/api-data.xlsx'
# sheet_name = 'Sheet2'
if __name__ == '__main__':
    read = ReadExcel(file_path, sheet_name)
    resul = read.read_data_obj()
    # resul = read.r_data_obj()

    for x in resul:
        print(x.__dict__)
        # print(TestBase().buildParam(x.__dict__['param']))
#
# if __name__ == '__main__':
#     test_da = DoExcel('E:/PycharmProjects/auto_test/test_data/api-data.xlsx', 'data-test').read_excel()
#
#     for itm in test_da:
#         print(itm)
#         # 返回的 值需要使用 eval()方法转换
#
#     test_da = DoExcel('E:/PycharmProjects/auto_test/test_data/api-data.xlsx', 'Sheet3')
#     test_da.write_back(1, 1, "python")
